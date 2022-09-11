#! python3
"""
------Ke软件电池数据自动填充程序V4.0------
------功能说明------
将同一文件夹下的所有测试结果的txt文件名以及对应的电池参数Voc、Jsc、FF、Eff四个参数填充到excel表格中。
------V2.0版本改进功能------
1、创建的文件名自动加上当天时间
2、输入的参数已经转换为float格式
(参数识别出来是string，要转换为float要在正则识别出数字之后转换，如果把含数字和空值的放在一起转换，空值无法转换为float，会报错)
3、改进了之前用参数列表赋值的方法，用嵌套字典直接对excel表格进行赋值，降低了空间复杂度。
(用嵌套字典进行赋值时，内部不用循环row，改成row在循环内自增，即可避免循环结束后表格中全都是最后一个文件的数据)
4、将输出数据按照文件名进行排序
(对字典按照键进行排序，用sorted+lambda会将字典变为元组，需要用强制类型转换 dict 将元组变回字典)
------V3.0版本改进功能------
1、添加tk可视化窗口，显示输入文件路径的文本框和一键提取的按钮
------V4.0版本改进功能------
1、优化tk可视化窗口外观
2、添加使用说明
3、重构代码，采用面向对象模块化编程
------V5.0版本改进功能------
1、提取Rs、Rsh数据，输出至excel表
------此版本存在的缺陷------
1、缺少单元格大小、对齐等美化处理
2、窗口仅满足了基本功能，缺乏设计性
3、未能考虑到错误条件和异常跳出的处理
4、当文件夹内有子文件夹时，错误
5、dataDic = dict(sorted(dataDic.items(), key=lambda x: int(x[0].split('-')[0])))
这句按照文件名排序，但是仅仅针对我的常用命名法，如1-1-1，当文件名中没有-时，就会出错。
实际上别人的文件名可能是任何组合，故而先把这句话注释掉，排序功能暂停。待升级。
------改进想法------
1、可供选择【按照某个参数进行排序】
2、可供选择【直接画出箱线图】
-------------备注-----------------
exe文件组装命令： pyinstaller -F FindertxtPortableV4.py -w
-w 作用是打开exe文件时不弹出命令行窗口
"""

import re  # 正则表达式模块
import os  # 文件夹操作模块
import datetime  # 时间模块
import openpyxl  # excel操作模块
from collections import defaultdict  # 默认字典
# tk窗口模块
from tkinter import *
from tkinter.filedialog import askdirectory
import tkinter.font as tkFont
import tkinter.messagebox

class MY_GUI():
    def __init__(self, windowName):  # 构造函数初始化窗口 和 文件路径
        self.windowName = windowName  # 传入一个窗口对象
        self.path = StringVar()  # 设个函数定义的变量会在entry中输入，一直追踪，可用get方法获取

    def selectPath(self):  # 实现用户选择路径
        path_ = askdirectory()  # 后面加下划线是为了避免与python的自带变量冲突
        self.path.set(path_)  # 在这里更新对象的 路径 变量path

    def error(self):  # 显示错误弹窗
        tkinter.messagebox.showinfo(title="提示", message="未选择路径")

    def setWindow(self):
        # 定制窗口的外观
        self.windowName["bg"] = "LightBlue"  # 窗口背景色
        self.windowName.attributes("-alpha", 0.95)  # 设置窗口不透明度 值越小透明度越高
        self.windowName.title("电池参数快捷统计工具")  # 窗口标题栏
        self.windowName.geometry('450x400+600+300')  # 窗口大小，350x350是窗口的大小，+600是距离左边距的距离，+300是距离上边距的距离
        #self.windowName.resizable(width=FALSE, height=FALSE)  # 拒绝用户调整窗口大小
        # 定制窗口的标签 xy控制标签的位置
        f = tkFont.Font(family='Microsoft YaHei UI', size=9, weight='bold')
        text = Text(self.windowName, width=53, height=6, bg="LightBlue", fg="Black", font=f)
        text.place(x=0, y=10)
        #text.pack()  # 加了pack 文本框就会跑到中间去
        text.insert(INSERT, "使用说明:\n")
        text.insert(INSERT, "1、将Ke2400S测试软件生成的txt文件放到同一文件夹内\n")
        text.insert(INSERT, "2、点击“选择电池文件夹”，选择该路径\n")
        text.insert(INSERT, "3、点击“一键导出”\n")
        text.insert(END, "4、程序将参数提取至该文件夹下自动生成的Excel表内\n")

        f1 = tkFont.Font(family='Microsoft YaHei UI', size=15, weight='bold')
        Button(self.windowName, text="选 择 电 池 文 件 夹",font=f1,command=self.selectPath,
               bg="Ivory").place(width=200,height=50,x=90, y=130)  # 调用选择路径的函数

        #f3 = tkFont.Font(family='Microsoft YaHei UI', size=12)
        f2 = tkFont.Font(family='Microsoft YaHei UI', size=10, weight='bold')
        Label(self.windowName, text="当前路径:",font=f2,bg="LightBlue", fg="DimGray").place(x=23, y=218)
        Entry(self.windowName, textvariable=self.path).place(width=200,height=25, x=108, y=215)  # 输入控件 显示文本内容。此处显示 路径属性 path

        f3 = tkFont.Font(family='Microsoft YaHei UI', size=15, weight='bold')
        Button(self.windowName, text="一 键 导 出",font=f3,
               bg="Ivory", command=self.option).place(width=200,height=50,x=90,y=273)  # 调用执行函数

        f4 = tkFont.Font(family='Microsoft YaHei UI', size=10, weight='bold')
        Label(self.windowName, text="designed by HangXu", font=f4,
              bg="LightBlue", fg="DimGray").place(x=20, y=350)

    def option(self):
        if self.path.get() == "":  # 文件路径为空时，报错
            self.error()
        else:  # 正确时，执行数据处理函数
            self.processPara(self.path.get())

    def getFileNums(self, inputPath):  # 统计该路径下的txt文件个数
        count = 0
        for filename in os.listdir(inputPath):
            if filename.endswith('.txt'):
                count += 1
        return count

    def processPara(self, inputPath):  # 主逻辑 处理数据
        # TODO: 建立待寻找参数的正则表达式。中文冒号可以直接匹配。名称、数字部分、单位分成三组。re.compile生成正则对象。后面再用search
        VocRegex = re.compile(r'(Voc：)([\d\.]{1,})(V)')
        JscRegex = re.compile(r'(Jsc：)([\d\.]{1,})(mA/cm\^2)')
        FFRegex = re.compile(r'(FF：)([\d\.]{1,})(%)')
        EffRegex = re.compile(r'(Eff：)([\d\.]{1,})(%)')
        RshRegex = re.compile(r'(Rsh：)([\d\.]{1,}[E\-\d]*)(Kohm)')
        RsRegex = re.compile(r'(Rs：)([\d\.]{1,}[E\-\d]*)(Kohm)')

        # TODO: 遍历文件夹中的txt数据文件，查找参数保存到字典中
        dataDic = defaultdict(dict)   # 用一个嵌套字典存放所有数据，父字典的key是文件名，value是子字典；子字典的key是参数名，value
        for filename in os.listdir(inputPath):  # os.listdir(路径) 返回一个列表，包含该路径下所有文件名
            if os.path.exists(os.path.join(inputPath, filename)) and filename.endswith('.txt'):  # 存在并且是txt文件
                testResult = open(os.path.join(inputPath, filename), encoding='gb18030', errors='ignore')
                # data是读取的文件参数数据
                data = testResult.read()
                testResult.close()
                if VocRegex.search(data): # search匹配data中第一个匹配到的
                    Voc = float(VocRegex.search(data).group(2))  # 正则第二组是数据
                else:
                    Voc = ''  # 匹配不到就置空
                if JscRegex.search(data):
                    Jsc = float(JscRegex.search(data).group(2))
                else:
                    Jsc = ''
                if FFRegex.search(data):
                    FF = float(FFRegex.search(data).group(2))
                else:
                    FF = ''
                if EffRegex.search(data):
                    Eff = float(EffRegex.search(data).group(2))
                else:
                    Eff = ''
                if RshRegex.search(data):
                    Rsh = float(RshRegex.search(data).group(2))
                else:
                    Rsh = ''
                if RsRegex.search(data):
                    Rs = float(RsRegex.search(data).group(2))
                else:
                    Rs = ''
                dataDic[filename] = {'Voc':Voc, 'Jsc':Jsc, 'FF':FF, 'Eff':Eff, 'Rsh':Rsh, 'Rs':Rs}  # 字典赋值

        # TODO: 创建并加载一个excel文件,更改工作表名
        nowTime = datetime.datetime.now().strftime('%Y-%m-%d')  # 获取当前日期
        os.chdir(inputPath)
        excelName = 'cellStats-' + nowTime + '.xlsx'  # 定义excel表格文件名
        wb = openpyxl.Workbook()
        wb.save(excelName)
        wb = openpyxl.load_workbook(excelName)
        sheet = wb.active  # 调用当前工作簿，用sheet.方法；保存工作表用wb.save
        sheet.title = "电池参数"  # 更改工作簿名

        # TODO: 输入表头
        titles = ['cell_num', 'Voc', 'Jsc', 'FF', 'Eff', 'Rsh', 'Rs']
        for i in range(0, 7):
            sheet.cell(row=1, column=i+1).value = titles[i]
        wb.save(excelName)

        # TODO: 直接把嵌套字典的内容输出至excel表格
        row = 2
        # 按文件名的第一个数字 也就是电池序号 进行排序
        # dataDic = dict(sorted(dataDic.items(), key=lambda x: int(x[0].split('-')[0])))  # 这句话会出导致文件名识别出问题
        for file, para in dataDic.items():
            # file 是主字典的键，表示当前电池序号。para是嵌套字典，表示当前电池的各项参数
            # 这里让row在循环内自增 即可避免表中所有的数都是最后一个文件的数据
            sheet.cell(row=row, column=1).value = file.rstrip('.txt')
            sheet.cell(row=row, column=2).value = para['Voc']
            sheet.cell(row=row, column=3).value = para['Jsc']
            sheet.cell(row=row, column=4).value = para['FF']
            sheet.cell(row=row, column=5).value = para['Eff']
            sheet.cell(row=row, column=6).value = para['Rsh']
            sheet.cell(row=row, column=7).value = para['Rs']
            row += 1
        wb.save(excelName)
        count = self.getFileNums(inputPath)
        tkinter.messagebox.showinfo(title="提示", message=str(count) + "个文件提取完成")  # 显示提示弹窗


if __name__ == "__main__":
    window1 = Tk()  # 创建一个窗口对象
    window1_GUI = MY_GUI(window1)  # 实例化
    window1_GUI.setWindow()
    window1.mainloop()